"""Bulk export/import of questions via Excel.

Three responsibilities:

  1. `build_template()` — emit a downloadable .xlsx with column headers,
     example rows, and data-validation dropdowns for the enum-shaped
     columns (difficulty, question_type, topic_code, domain) so admins
     can't typo them.

  2. `build_export(rows)` — emit the SAME sheet shape, but pre-filled
     with every existing question (id + all fields + exam-set memberships
     + options). This is what "Download" gives an admin: the live data,
     ready to edit and re-upload.

  3. `parse_workbook(stream)` — read an uploaded sheet, map each row to a
     `QuestionAdminIn` payload plus its `id` (blank = create, present =
     update) and `exam_sets` slug list. Caller (the endpoint) decides how
     to apply them.

Format (wide, one row per question):

    id                  (existing id → update; blank or unknown id → create new)
    stem                (required)
    topic_code          (required, case-insensitive: BU/DU/DP/MD/EV/DE)
    difficulty          (required: easy / medium / hard)
    question_type       (default single_choice; multi_choice allowed)
    domain              (ECO domain code: D-I … D-V; blank = unassigned)
    task                (optional)
    enablers            (optional, comma-separated → list[str])
    remarks             (optional, admin-only note)
    explanation         (optional, shown to learner after submit)
    is_active           (default true; accepts y/n, true/false, 1/0)
    exam_sets           (comma-separated set slugs — AUTHORITATIVE on import:
                         the question's memberships are synced to exactly
                         this list; clearing the cell removes it from all sets)
    option_a_text       (required) … option_f_* (optional)

Hard caps (enforced at the endpoint, not here):
    file size  ≤ 5 MB
    row count  ≤ 500 questions per upload
"""
from io import BytesIO
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from pydantic import ValidationError as PydanticValidationError

from app.core import domains as domain_registry
from app.schemas.question import QuestionAdminIn, QuestionOptionIn


# Column layout — must match the docstring above. Single source of truth
# for both the template writer and the parser.
COLUMNS: list[tuple[str, str]] = [
    ("id",             "Question id. An existing id UPDATES that question; "
                       "blank — or an unknown id — CREATES a new one."),
    ("stem",           "Question stem (required)"),
    ("topic_code",     "CPMAI phase code: BU, DU, DP, MD, EV, or DE (required)"),
    ("difficulty",     "easy / medium / hard (required)"),
    ("question_type",  "single_choice (default) or multi_choice"),
    ("domain",         "ECO domain code: D-I … D-V (blank = unassigned)"),
    ("task",           "Optional task description"),
    ("enablers",       "Optional comma-separated list of enablers"),
    ("remarks",        "Optional admin-only note (not shown to learner)"),
    ("explanation",    "Optional general explanation, shown after submit"),
    ("is_active",      "true (default) / false / 1 / 0 / y / n"),
    ("exam_sets",      "Comma-separated exam-set slugs. AUTHORITATIVE on "
                       "upload: memberships are synced to exactly this list "
                       "(clear the cell to remove from all sets)."),
]
OPTION_LETTERS = ("A", "B", "C", "D", "E", "F")
for L in OPTION_LETTERS:
    COLUMNS += [
        (f"option_{L.lower()}_text",      f"Option {L} text"
            + (" (required)" if L in ("A", "B") else " (leave blank if unused)")),
        (f"option_{L.lower()}_is_correct", f"Option {L} correctness: true/false"),
        (f"option_{L.lower()}_reasoning",  f"Option {L} reasoning (correct → why; "
                                            f"wrong → why wrong)"),
    ]
HEADERS = [c[0] for c in COLUMNS]
HEADER_NOTES = {c[0]: c[1] for c in COLUMNS}

# Comma-separated domain codes for the dropdown + error messages.
_DOMAIN_CODES = [d.code for d in domain_registry.all_domains()]


# --------------------------------------------------------- bool parsing
_TRUTHY = {"true", "t", "yes", "y", "1"}
_FALSY  = {"false", "f", "no",  "n", "0"}


def _parse_bool(v, default: bool | None = None) -> bool:
    """Forgiving bool parser — Excel cells come through as int / str / bool."""
    if v is None or (isinstance(v, str) and not v.strip()):
        if default is None:
            raise ValueError("expected a yes/no value, got blank")
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in _TRUTHY: return True
    if s in _FALSY:  return False
    raise ValueError(f"could not interpret {v!r} as yes/no")


def _cell_str(v) -> str:
    """Cell → trimmed string. Empty cells become ''."""
    if v is None:
        return ""
    return str(v).strip()


# =========================================================== template
# Three example rows — pre-filled so admins can copy-and-modify a blank sheet.
_EXAMPLE_ROWS: list[dict] = [
    {
        "stem": "Which CPMAI phase is dedicated to assessing data quality?",
        "topic_code": "DU", "difficulty": "easy",
        "question_type": "single_choice",
        "domain": "D-III",
        "explanation": "Phase 2 (Data Understanding) is where data is profiled.",
        "is_active": "true",
        "option_a_text": "Phase 1 — Business Understanding", "option_a_is_correct": "false",
        "option_a_reasoning": "Phase 1 defines the business goal, not data quality.",
        "option_b_text": "Phase 2 — Data Understanding", "option_b_is_correct": "true",
        "option_b_reasoning": "Correct — Phase 2 is dedicated to data assessment.",
        "option_c_text": "Phase 4 — Modeling", "option_c_is_correct": "false",
        "option_c_reasoning": "Modeling consumes prepared data; quality is assessed earlier.",
        "option_d_text": "Phase 6 — Operationalization", "option_d_is_correct": "false",
        "option_d_reasoning": "Operationalization is for deployed models, not raw data.",
    },
    {
        "stem": "Which phases involve hands-on work with data? (pick all that apply)",
        "topic_code": "DP", "difficulty": "medium",
        "question_type": "multi_choice",
        "domain": "D-III",
        "is_active": "true",
        "option_a_text": "Data Understanding", "option_a_is_correct": "true",
        "option_a_reasoning": "Data profiling and quality assessment happen here.",
        "option_b_text": "Data Preparation", "option_b_is_correct": "true",
        "option_b_reasoning": "Cleansing, transformation, and feature engineering.",
        "option_c_text": "Modeling", "option_c_is_correct": "false",
        "option_c_reasoning": "Modeling consumes data but doesn't manipulate it directly.",
        "option_d_text": "Business Understanding", "option_d_is_correct": "false",
        "option_d_reasoning": "Business Understanding is goal-setting, not data work.",
    },
    {
        "stem": "Minimal example: what does CPMAI stand for?",
        "topic_code": "BU", "difficulty": "easy", "domain": "D-II",
        "option_a_text": "Cognitive Project Management for AI", "option_a_is_correct": "true",
        "option_b_text": "Continuous Process Modeling and Iteration", "option_b_is_correct": "false",
    },
]


def _write_workbook(data_rows: list[dict]) -> bytes:
    """Write the canonical sheet (header + validations + the given rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    header_fill = PatternFill("solid", fgColor="E5E7EB")
    header_font = Font(bold=True)

    # Header row.
    for i, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.comment = Comment(HEADER_NOTES[name], "CPMAI Bulk Upload")

    ws.freeze_panes = "A2"
    # Column widths — guesstimated for readability; admins can resize.
    for i, name in enumerate(HEADERS, start=1):
        if name == "stem" or name == "explanation":
            ws.column_dimensions[get_column_letter(i)].width = 60
        elif name.endswith("_text") or name.endswith("_reasoning") or name == "exam_sets":
            ws.column_dimensions[get_column_letter(i)].width = 35
        else:
            ws.column_dimensions[get_column_letter(i)].width = 16

    # Data-validation dropdowns. Apply to a generous range so admins
    # don't have to extend the validation when adding rows.

    def col(name: str) -> str:
        return get_column_letter(HEADERS.index(name) + 1)

    def col_range(name: str, start_row: int = 2, end_row: int = 2000) -> str:
        c = col(name)
        return f"{c}{start_row}:{c}{end_row}"

    diff_dv = DataValidation(type="list", formula1='"easy,medium,hard"',
                              allow_blank=False)
    diff_dv.error = "Use easy / medium / hard"
    diff_dv.errorTitle = "Invalid difficulty"
    ws.add_data_validation(diff_dv)
    diff_dv.add(col_range('difficulty'))

    qtype_dv = DataValidation(type="list",
                               formula1='"single_choice,multi_choice"',
                               allow_blank=True)
    qtype_dv.error = "Use single_choice or multi_choice"
    qtype_dv.errorTitle = "Invalid question_type"
    ws.add_data_validation(qtype_dv)
    qtype_dv.add(col_range('question_type'))

    bool_dv = DataValidation(type="list", formula1='"true,false"',
                              allow_blank=True)
    bool_dv.error = "Use true / false"
    bool_dv.errorTitle = "Invalid boolean"
    ws.add_data_validation(bool_dv)
    bool_dv.add(col_range('is_active'))
    for L in OPTION_LETTERS:
        bool_dv.add(col_range(f'option_{L.lower()}_is_correct'))

    topic_dv = DataValidation(type="list",
                               formula1='"BU,DU,DP,MD,EV,DE"',
                               allow_blank=False)
    topic_dv.error = "Use one of: BU, DU, DP, MD, EV, DE"
    topic_dv.errorTitle = "Invalid topic_code"
    ws.add_data_validation(topic_dv)
    topic_dv.add(col_range('topic_code'))

    domain_dv = DataValidation(type="list",
                                formula1='"' + ",".join(_DOMAIN_CODES) + '"',
                                allow_blank=True)
    domain_dv.error = "Use one of: " + ", ".join(_DOMAIN_CODES)
    domain_dv.errorTitle = "Invalid domain"
    ws.add_data_validation(domain_dv)
    domain_dv.add(col_range('domain'))

    for rownum, ex in enumerate(data_rows, start=2):
        for i, name in enumerate(HEADERS, start=1):
            if name in ex and ex[name] is not None:
                ws.cell(row=rownum, column=i, value=ex[name])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template() -> bytes:
    """Generate the blank .xlsx admins download (header + 3 example rows
    + dropdown validations). Used when there's nothing to export yet."""
    return _write_workbook(_EXAMPLE_ROWS)


def build_export(rows: list[dict]) -> bytes:
    """Generate an .xlsx pre-filled with existing questions (one dict per
    question, keys matching HEADERS). Falls back to the example rows when
    the bank is empty so the admin still gets a usable starting sheet."""
    return _write_workbook(rows if rows else _EXAMPLE_ROWS)


def question_to_row(q, topic_code: str, set_slugs: list[str]) -> dict:
    """Flatten a Question ORM object + its set memberships into a row dict
    for `build_export`. `topic_code` is supplied by the caller (the Question
    model has no `topic` relationship)."""
    row: dict = {
        "id": q.id,
        "stem": q.stem,
        "topic_code": topic_code,
        "difficulty": q.difficulty.value if hasattr(q.difficulty, "value") else q.difficulty,
        "question_type": (q.question_type.value
                          if hasattr(q.question_type, "value") else q.question_type),
        "domain": q.domain or "",
        "task": q.task or "",
        "enablers": ", ".join(q.enablers or []),
        "remarks": q.remarks or "",
        "explanation": q.explanation or "",
        "is_active": "true" if q.is_active else "false",
        "exam_sets": ", ".join(set_slugs),
    }
    for opt in q.options:
        L = opt.option_letter.lower()
        row[f"option_{L}_text"] = opt.text
        row[f"option_{L}_is_correct"] = "true" if opt.is_correct else "false"
        row[f"option_{L}_reasoning"] = opt.reasoning or ""
    return row


# ============================================================ parser
@dataclass
class ParsedRow:
    """One successfully-parsed upload row."""
    row_num: int
    payload: QuestionAdminIn
    topic_code: str
    question_id: int | None       # None → create; int → update existing
    # Authoritative membership list when the `exam_sets` column is present;
    # None when the column is absent entirely (→ leave memberships untouched,
    # so an old sheet without the column never wipes associations).
    set_slugs: list[str] | None


@dataclass
class ParseResult:
    """Outcome of parsing the upload. The endpoint applies `valid` entries
    and reports `errors` so the admin can fix failing rows and re-upload."""
    valid: list[ParsedRow]
    errors: list[dict]            # {row, field, message}


def _build_row(row: dict, row_num: int, *, sets_present: bool) -> ParsedRow:
    """Validate + coerce a single sheet row. Raises ValueError with a
    human-readable message on any problem. topic_code is resolved to a
    topic_id by the endpoint (no DB session here)."""
    stem = _cell_str(row.get("stem"))
    if not stem:
        raise ValueError("stem is required")

    # id: blank → create; otherwise must be a positive integer.
    question_id: int | None = None
    id_raw = row.get("id")
    if id_raw is not None and _cell_str(id_raw) != "":
        try:
            question_id = int(float(id_raw))   # tolerate "12" / 12 / 12.0
        except (TypeError, ValueError):
            raise ValueError(f"id must be a whole number or blank, got {id_raw!r}")
        if question_id <= 0:
            raise ValueError(f"id must be positive, got {question_id}")

    topic_code = _cell_str(row.get("topic_code")).upper()
    if not topic_code:
        raise ValueError("topic_code is required")

    difficulty = _cell_str(row.get("difficulty")).lower()
    if difficulty not in {"easy", "medium", "hard"}:
        raise ValueError(f"difficulty must be easy/medium/hard, got {difficulty!r}")

    qt = _cell_str(row.get("question_type")).lower() or "single_choice"
    if qt not in {"single_choice", "multi_choice"}:
        raise ValueError(f"question_type must be single_choice or multi_choice, got {qt!r}")

    # domain: blank allowed. A recognised value (code/name/slug) is
    # normalised to its canonical code; an unrecognised value is preserved
    # as-is. Tolerating legacy free-text keeps an export → re-import of
    # pre-existing data a clean no-op rather than a wall of row errors.
    domain_raw = _cell_str(row.get("domain"))
    domain_value: str | None = None
    if domain_raw:
        d = domain_registry.get(domain_raw)
        domain_value = d.code if d else domain_raw

    enablers_raw = _cell_str(row.get("enablers"))
    enablers = [e.strip() for e in enablers_raw.split(",") if e.strip()] \
                if enablers_raw else []

    # None when the column is absent (don't touch memberships); a (possibly
    # empty) list when present (authoritative).
    if sets_present:
        sets_raw = _cell_str(row.get("exam_sets"))
        set_slugs: list[str] | None = [
            s.strip() for s in sets_raw.split(",") if s.strip()]
    else:
        set_slugs = None

    is_active = _parse_bool(row.get("is_active"), default=True)

    options: list[QuestionOptionIn] = []
    for L in OPTION_LETTERS:
        text = _cell_str(row.get(f"option_{L.lower()}_text"))
        if not text:
            continue   # blank option columns are skipped
        try:
            is_correct = _parse_bool(row.get(f"option_{L.lower()}_is_correct"),
                                      default=False)
        except ValueError as e:
            raise ValueError(f"option_{L.lower()}_is_correct: {e}")
        reasoning = _cell_str(row.get(f"option_{L.lower()}_reasoning")) or None
        options.append(QuestionOptionIn(
            option_letter=L, text=text,
            is_correct=is_correct, reasoning=reasoning,
        ))

    if len(options) < 2:
        raise ValueError("at least 2 options required (option_a_text + option_b_text)")

    # topic_id is set later by the caller after looking up topic_code.
    # Use 0 as a placeholder — Pydantic validates type only.
    try:
        payload = QuestionAdminIn(
            stem=stem, topic_id=0,
            domain=domain_value,
            task=(_cell_str(row.get("task")) or None),
            enablers=enablers,
            remarks=(_cell_str(row.get("remarks")) or None),
            difficulty=difficulty,                 # type: ignore[arg-type]
            question_type=qt,                      # type: ignore[arg-type]
            explanation=(_cell_str(row.get("explanation")) or None),
            options=options,
            is_active=is_active,
        )
    except PydanticValidationError as e:
        raise ValueError(f"schema validation: {e.errors()[0]['msg']}")
    return ParsedRow(row_num=row_num, payload=payload, topic_code=topic_code,
                     question_id=question_id, set_slugs=set_slugs)


def parse_workbook(stream: bytes, *, max_rows: int = 500) -> ParseResult:
    """Read a .xlsx upload and return per-row parsed rows + errors.

    `max_rows` is enforced ABOVE the schema: a too-big sheet returns an
    immediate "too many" error before any row is parsed.
    """
    try:
        wb = load_workbook(BytesIO(stream), data_only=True, read_only=True)
    except Exception as e:
        return ParseResult(valid=[], errors=[{
            "row": 0, "field": "file",
            "message": f"could not open as .xlsx: {e}"}])

    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return ParseResult(valid=[], errors=[{
            "row": 0, "field": "file", "message": "sheet is empty"}])

    headers_in_file = [_cell_str(h) for h in header_row]
    # `id` and `exam_sets` are optional columns for back-compat with older
    # sheets — everything else must be present.
    optional = {"id", "exam_sets"}
    missing = [h for h in HEADERS if h not in headers_in_file and h not in optional]
    if missing:
        return ParseResult(valid=[], errors=[{
            "row": 1, "field": "headers",
            "message": (f"missing required header columns: {missing}. "
                         f"Download the latest template via the "
                         f"'Download' button.")}])

    name_at = {i: name for i, name in enumerate(headers_in_file)}
    sets_present = "exam_sets" in headers_in_file

    valid: list[ParsedRow] = []
    errors: list[dict] = []
    seen_data_rows = 0

    for row_num, row_values in enumerate(rows_iter, start=2):
        if all(v is None or (isinstance(v, str) and not v.strip())
                for v in row_values):
            continue
        seen_data_rows += 1
        if seen_data_rows > max_rows:
            errors.append({
                "row": row_num, "field": "file",
                "message": (f"too many rows: capped at {max_rows} per upload. "
                             f"Split into smaller files."),
            })
            break
        row = {name_at.get(i): v for i, v in enumerate(row_values)}
        try:
            valid.append(_build_row(row, row_num, sets_present=sets_present))
        except ValueError as e:
            errors.append({
                "row": row_num, "field": "row",
                "message": str(e),
            })

    return ParseResult(valid=valid, errors=errors)
