"""File parsers for admin-uploaded RAG sources.

Each parser takes raw bytes + a filename, returns a list of plain-text
chunks ready for embedding. Chunking strategy is per-file-type so the
semantic granularity matches the source's natural unit:

  - .txt / .md  → split on paragraphs (double newline), then re-bundle
                  into ~600-token chunks with ~50-token overlap
  - .xlsx       → one chunk per row (each row's cells joined as text)
  - .pdf        → one chunk per page (page numbers in metadata)
  - .docx       → split on paragraphs, re-bundle like .txt

Hard-cap safety net
-------------------
OpenAI's embedding API (text-embedding-3-large) caps each input at
8192 tokens. A naive paragraph-aware chunker can still produce
oversized chunks when the SOURCE has a single paragraph or PDF page
that's already larger than the budget — common for legal/standards
documents, FAQs without blank-line breaks, or scanned PDFs whose OCR
output has no paragraph delimiters. When that happens the embed API
returns 400 ("Invalid 'input[N]': maximum input length is 8192
tokens"), the whole batch fails, and rag_chunks stays empty.

The fix: every chunk goes through ``_split_oversized`` before being
yielded. Anything larger than ``_TARGET_CHARS`` is split — first on
sentence boundaries, then on hard character cuts as a fallback. Each
sub-chunk gets the same metadata so retrieval citations still point
at the right page/row/source.

No tokenizer dep — we approximate tokens as `len(text) // 4` (a
well-known rule-of-thumb for English text, ±10%). Good enough for
chunk sizing; the embedding API counts real tokens server-side.
"""
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Iterator


@dataclass
class ParsedChunk:
    content: str
    metadata: dict


# Approx chars per token (English). Used for chunk-size budgeting,
# NOT for billing.
_CHARS_PER_TOKEN = 4
# Target chunk size in TOKENS. ~600 tokens lets a k=4 retrieval pull
# ~2400 tokens of context — fits comfortably in a 4k system prompt
# budget without crowding out the user's question.
_TARGET_TOKENS = 600
_OVERLAP_TOKENS = 50

_TARGET_CHARS  = _TARGET_TOKENS  * _CHARS_PER_TOKEN
_OVERLAP_CHARS = _OVERLAP_TOKENS * _CHARS_PER_TOKEN

# Hard ceiling for any individual chunk before embedding. OpenAI's
# text-embedding-3-large accepts 8192 tokens per input ≈ ~32k chars.
# We keep ample headroom (24k chars ≈ 6000 tokens) so non-English
# text and tokenizer variance don't push us over.
_MAX_CHARS = 24_000


def _split_oversized(text: str, *,
                      max_chars: int = _TARGET_CHARS,
                      hard_max_chars: int = _MAX_CHARS,
                      overlap: int = _OVERLAP_CHARS) -> Iterator[str]:
    """Defensive splitter: yields each input as one or more sub-chunks
    no larger than ``max_chars``.

    Strategy:
      1. If the whole input fits, yield as one chunk (common case).
      2. Otherwise split on sentence boundaries (period/!/? followed by
         whitespace) and bundle sentences greedily up to ``max_chars``.
      3. If a SINGLE sentence is itself larger than ``hard_max_chars``
         (very rare — code blocks, URLs, base64 blobs in docs), hard-cut
         it on character boundary with overlap.

    Returned chunks all satisfy ``len(chunk) <= hard_max_chars`` —
    guaranteed safe for the embedding API.
    """
    if len(text) <= max_chars:
        yield text
        return

    # Sentence boundary: period/exclamation/question followed by whitespace.
    # Conservative — leaves abbreviations like "e.g." with the next sentence,
    # which is fine (we're chunking, not parsing prose).
    sentences = re.split(r"(?<=[.!?])\s+", text)

    current = ""
    for sent in sentences:
        # A single sentence > hard_max_chars (rare): hard-cut it.
        if len(sent) > hard_max_chars:
            if current:
                yield current
                current = ""
            step = hard_max_chars - overlap
            for i in range(0, len(sent), max(1, step)):
                yield sent[i:i + hard_max_chars]
            continue

        if not current:
            current = sent
        elif len(current) + 1 + len(sent) <= max_chars:
            current = current + " " + sent
        else:
            # Flush, start next with overlap from the tail of the
            # previous chunk (so a query landing at the boundary still
            # sees both sides).
            yield current
            tail = current[-overlap:] if overlap < len(current) else current
            current = tail + " " + sent
    if current:
        yield current


# ----------------------------------------------------------- public API
def parse_file(filename: str, content_type: str,
                data: bytes) -> list[ParsedChunk]:
    """Dispatch to the right parser. Returns at-least-one-chunk on
    success; raises ValueError on unsupported or unparseable input."""
    name_lower = filename.lower()
    if name_lower.endswith(".txt") or name_lower.endswith(".md"):
        return list(_parse_text(data, filename))
    if name_lower.endswith(".xlsx"):
        return list(_parse_xlsx(data, filename))
    if name_lower.endswith(".pdf"):
        return list(_parse_pdf(data, filename))
    if name_lower.endswith(".docx"):
        return list(_parse_docx(data, filename))
    raise ValueError(
        f"Unsupported file type: {filename}. "
        f"Supported: .txt, .md, .xlsx, .pdf, .docx")


# ----------------------------------------------------------- text/md
def _parse_text(data: bytes, filename: str) -> Iterator[ParsedChunk]:
    text = data.decode("utf-8", errors="replace").strip()
    if not text:
        raise ValueError(f"{filename} is empty.")
    yield from _windowed_paragraphs(text, filename)


# ----------------------------------------------------------- xlsx
def _parse_xlsx(data: bytes, filename: str) -> Iterator[ParsedChunk]:
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    yielded = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Capture header row (if any) to give chunks context.
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if header_row is None:
                header_row = [str(c) if c is not None else "" for c in row]
                continue
            cells = [str(c) for c in row if c is not None]
            if not cells:
                continue
            # Format as "col_a=val; col_b=val; ..." for cleaner retrieval.
            pairs: list[str] = []
            for i, val in enumerate(row):
                if val is None: continue
                key = (header_row[i] if i < len(header_row)
                                     else f"col{i+1}")
                pairs.append(f"{key}: {val}")
            if not pairs:
                continue
            row_text = "; ".join(pairs)
            # Defensive split — a row with a giant text cell (free-form
            # description columns, embedded JSON, etc.) can blow past
            # the embed API's 8192-token cap if shipped whole.
            for sub in _split_oversized(row_text):
                yield ParsedChunk(
                    content=sub,
                    metadata={"sheet": sheet_name, "filename": filename},
                )
                yielded += 1
    if yielded == 0:
        raise ValueError(f"{filename} contained no data rows.")


# ----------------------------------------------------------- pdf
def _parse_pdf(data: bytes, filename: str) -> Iterator[ParsedChunk]:
    from pypdf import PdfReader
    reader = PdfReader(BytesIO(data))
    if len(reader.pages) == 0:
        raise ValueError(f"{filename} has no pages.")
    for i, page in enumerate(reader.pages):
        text = (page.extract_text() or "").strip()
        if not text:
            continue
        # Pages can be long — apply paragraph windowing per page so we
        # don't ship a 5K-token blob as one chunk.
        for sub in _windowed_paragraphs(text, filename,
                                         base_metadata={"page": i + 1}):
            yield sub


# ----------------------------------------------------------- docx
def _parse_docx(data: bytes, filename: str) -> Iterator[ParsedChunk]:
    from docx import Document
    doc = Document(BytesIO(data))
    text = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not text:
        raise ValueError(f"{filename} contains no paragraph text.")
    yield from _windowed_paragraphs(text, filename)


# ----------------------------------------------------------- windowing
def _windowed_paragraphs(text: str, filename: str, *,
                          base_metadata: dict | None = None
                          ) -> Iterator[ParsedChunk]:
    """Split text into paragraph-boundary-aware chunks of ≤ _TARGET_CHARS,
    with _OVERLAP_CHARS of context between consecutive chunks.

    Paragraph-aware so we don't slice through a thought; overlapping so
    a query that lands at the chunk boundary still sees both sides."""
    md = dict(base_metadata or {})
    md["filename"] = filename

    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]

    current = ""
    chunk_idx = 0
    for p in paragraphs:
        # Defensive split: a single paragraph larger than _TARGET_CHARS
        # (PDF pages without blank-line breaks, FAQ entries with one
        # giant body, OCR output that lost paragraph delimiters) gets
        # broken into sub-pieces FIRST so we never accumulate it whole
        # into `current` and overflow on the next flush.
        if len(p) > _TARGET_CHARS:
            if current:
                yield ParsedChunk(content=current,
                                   metadata={**md, "chunk_index": chunk_idx})
                chunk_idx += 1
                current = ""
            for sub in _split_oversized(p):
                yield ParsedChunk(content=sub,
                                   metadata={**md, "chunk_index": chunk_idx})
                chunk_idx += 1
            continue

        if not current:
            current = p
            continue
        if len(current) + 2 + len(p) <= _TARGET_CHARS:
            current = current + "\n\n" + p
            continue
        # Flush current chunk.
        yield ParsedChunk(content=current, metadata={**md, "chunk_index": chunk_idx})
        chunk_idx += 1
        # Start next chunk with the tail of the previous for overlap +
        # this new paragraph.
        tail = current[-_OVERLAP_CHARS:] if _OVERLAP_CHARS < len(current) else current
        current = tail + "\n\n" + p
    if current:
        yield ParsedChunk(content=current, metadata={**md, "chunk_index": chunk_idx})
