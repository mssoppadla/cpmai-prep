"""Bulk export → edit → re-import round-trip, including exam-set
association sync (the headline requirement: download includes all live data,
upload reflects every change including memberships)."""
from io import BytesIO

from openpyxl import load_workbook

from app.models.exam_set import ExamSet, ExamSetQuestion
from app.models.question import Question, QuestionOption, Difficulty
from app.models.topic import Topic
from tests.conftest import auth_header


def _seed_question(db, *, topic="BU", domain="D-II", stem="Seed question") -> Question:
    t = db.query(Topic).filter_by(code=topic).first()
    q = Question(stem=stem + " " + "x" * 15, topic_id=t.id, domain=domain,
                 difficulty=Difficulty.EASY, is_active=True)
    q.options = [
        QuestionOption(option_letter="A", text="wrong", is_correct=False),
        QuestionOption(option_letter="B", text="right", is_correct=True),
    ]
    db.add(q); db.commit(); db.refresh(q)
    return q


def _empty_set(db, admin, slug="s1") -> ExamSet:
    es = ExamSet(name=slug.upper(), slug=slug, time_limit_minutes=30,
                 passing_score=70, is_active=True, created_by=admin.id)
    db.add(es); db.commit(); db.refresh(es)
    return es


def _export(client, headers) -> "Worksheet":
    r = client.get("/api/v1/admin/questions/export", headers=headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    return wb.active


def _headers_index(ws) -> dict:
    return {c.value: i for i, c in enumerate(ws[1])}  # name → 0-based col


def _row_for_id(ws, idx, qid) -> int:
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=idx["id"] + 1).value == qid:
            return r
    raise AssertionError(f"question id {qid} not found in export")


def _upload(client, headers, ws) -> dict:
    buf = BytesIO()
    ws.parent.save(buf)
    buf.seek(0)
    r = client.post("/api/v1/admin/questions/bulk-upload", headers=headers,
                    files={"file": ("q.xlsx", buf.read(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    return r.json()


def test_export_import_updates_and_syncs_memberships(client, admin, db):
    headers = auth_header(client, admin.email)
    q1 = _seed_question(db, domain="D-II", stem="Round trip Q1")
    s1 = _empty_set(db, admin, "s1")

    # 1) EXPORT — the live row is present, with empty exam_sets.
    ws = _export(client, headers)
    idx = _headers_index(ws)
    r1 = _row_for_id(ws, idx, q1.id)
    assert ws.cell(row=r1, column=idx["domain"] + 1).value == "D-II"
    assert (ws.cell(row=r1, column=idx["exam_sets"] + 1).value or "") == ""

    # 2) EDIT — change Q1's domain + tag it into s1, and append a NEW row.
    ws.cell(row=r1, column=idx["domain"] + 1, value="D-IV")
    ws.cell(row=r1, column=idx["exam_sets"] + 1, value="s1")
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=idx["stem"] + 1, value="Brand new question " + "y" * 12)
    ws.cell(row=new_row, column=idx["topic_code"] + 1, value="BU")
    ws.cell(row=new_row, column=idx["difficulty"] + 1, value="easy")
    ws.cell(row=new_row, column=idx["domain"] + 1, value="D-I")
    ws.cell(row=new_row, column=idx["exam_sets"] + 1, value="s1")
    ws.cell(row=new_row, column=idx["option_a_text"] + 1, value="opt a")
    ws.cell(row=new_row, column=idx["option_a_is_correct"] + 1, value="false")
    ws.cell(row=new_row, column=idx["option_b_text"] + 1, value="opt b")
    ws.cell(row=new_row, column=idx["option_b_is_correct"] + 1, value="true")

    # 3) UPLOAD — one update (Q1), one create (new row).
    res = _upload(client, headers, ws)
    assert res["updated"] == 1, res
    assert res["created"] == 1, res
    assert res["errors"] == [], res

    db.expire_all()
    q1 = db.get(Question, q1.id)
    assert q1.domain == "D-IV"   # field updated in place
    # Both questions are now members of s1.
    members = {link.question_id for link in
               db.query(ExamSetQuestion).filter_by(exam_set_id=s1.id).all()}
    assert q1.id in members
    assert len(members) == 2


def test_export_import_removes_membership_when_cleared(client, admin, db):
    headers = auth_header(client, admin.email)
    q1 = _seed_question(db, stem="Membership clear Q")
    s1 = _empty_set(db, admin, "s1")
    db.add(ExamSetQuestion(exam_set_id=s1.id, question_id=q1.id,
                           position=0, added_by=admin.id))
    db.commit()

    ws = _export(client, headers)
    idx = _headers_index(ws)
    r1 = _row_for_id(ws, idx, q1.id)
    # Export should show the current membership.
    assert ws.cell(row=r1, column=idx["exam_sets"] + 1).value == "s1"

    # Clear the cell → authoritative sync removes it from the set. (Assign
    # .value directly: ws.cell(..., value=None) is a no-op in openpyxl.)
    ws.cell(row=r1, column=idx["exam_sets"] + 1).value = None
    res = _upload(client, headers, ws)
    assert res["updated"] == 1, res

    remaining = db.query(ExamSetQuestion).filter_by(exam_set_id=s1.id).count()
    assert remaining == 0


def test_import_unknown_set_slug_is_row_error(client, admin, db):
    headers = auth_header(client, admin.email)
    q1 = _seed_question(db, stem="Bad slug Q")

    ws = _export(client, headers)
    idx = _headers_index(ws)
    r1 = _row_for_id(ws, idx, q1.id)
    ws.cell(row=r1, column=idx["exam_sets"] + 1, value="does-not-exist")
    res = _upload(client, headers, ws)
    assert res["updated"] == 0, res
    assert len(res["errors"]) == 1
    assert "unknown exam_set slug" in res["errors"][0]["message"]


def test_import_normalizes_known_domain_name(client, admin, db):
    """A recognised domain *name* (not just the code) normalises to the
    canonical code on import."""
    headers = auth_header(client, admin.email)
    q1 = _seed_question(db, domain="D-II", stem="Name normalize Q")

    ws = _export(client, headers)
    idx = _headers_index(ws)
    r1 = _row_for_id(ws, idx, q1.id)
    ws.cell(row=r1, column=idx["domain"] + 1, value="Trustworthy AI")
    res = _upload(client, headers, ws)
    assert res["updated"] == 1, res
    assert res["errors"] == []
    db.expire_all()
    assert db.get(Question, q1.id).domain == "D-I"


def test_import_preserves_legacy_freetext_domain(client, admin, db):
    """Backward compatibility: an unrecognised (legacy free-text) domain is
    preserved as-is rather than erroring, so re-importing pre-existing data
    is a clean no-op."""
    headers = auth_header(client, admin.email)
    # Seed a row that still holds an old free-text domain value.
    q1 = _seed_question(db, domain="Data Understanding > Quality",
                        stem="Legacy domain Q")

    ws = _export(client, headers)
    idx = _headers_index(ws)
    r1 = _row_for_id(ws, idx, q1.id)
    # Export carried the legacy value through unchanged.
    assert ws.cell(row=r1, column=idx["domain"] + 1).value == "Data Understanding > Quality"
    res = _upload(client, headers, ws)
    assert res["errors"] == [], res
    db.expire_all()
    assert db.get(Question, q1.id).domain == "Data Understanding > Quality"


def test_import_unknown_id_creates_new_question(client, admin, db):
    """An id that doesn't exist (e.g. a new row an admin hand-numbered) is
    upserted as a CREATE, not rejected — and the DB assigns the real id."""
    headers = auth_header(client, admin.email)
    _seed_question(db, stem="Existing Q for upsert")

    ws = _export(client, headers)
    idx = _headers_index(ws)
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=idx["id"] + 1, value=99999)          # non-existent id
    ws.cell(row=new_row, column=idx["stem"] + 1, value="New via unknown id " + "z" * 12)
    ws.cell(row=new_row, column=idx["topic_code"] + 1, value="BU")
    ws.cell(row=new_row, column=idx["difficulty"] + 1, value="easy")
    ws.cell(row=new_row, column=idx["option_a_text"] + 1, value="a")
    ws.cell(row=new_row, column=idx["option_a_is_correct"] + 1, value="false")
    ws.cell(row=new_row, column=idx["option_b_text"] + 1, value="b")
    ws.cell(row=new_row, column=idx["option_b_is_correct"] + 1, value="true")

    res = _upload(client, headers, ws)
    assert res["created"] == 1, res     # the unknown-id row created a new question
    assert res["errors"] == [], res
    # The sheet's id is ignored — nothing literally lands at id 99999.
    assert db.get(Question, 99999) is None
    created = db.query(Question).filter(
        Question.stem.like("New via unknown id%")).one()
    assert created.id != 99999
