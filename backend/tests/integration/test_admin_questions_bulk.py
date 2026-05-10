"""Bulk-upload questions admin endpoint.

Two endpoints under test:
  GET  /admin/questions/bulk-template   — downloads the .xlsx template
  POST /admin/questions/bulk-upload     — parses + commits valid rows

The interesting behaviours:
  - Per-row partial success: a sheet with 3 valid + 2 broken rows
    creates 3 questions and returns 2 errors. Bad rows must NOT
    poison the good ones.
  - The bulk path runs the SAME `_validate(payload)` function the
    single-question POST uses. We assert this by sending the same
    rule violations and expecting the same error messages.
  - File-shape guards (too large, empty, missing headers, unknown
    topic_code) all return clean per-row errors instead of 500s.
  - Audit log captures the bulk action so an admin's "what did I just
    do?" question has a paper trail.
"""
import io
import pytest
from openpyxl import Workbook, load_workbook

from app.api.v1.endpoints.admin import questions as questions_ep
from app.models.audit_log import AuditLog
from app.models.question import Question
from app.services.question_excel import HEADERS, build_template
from tests.conftest import auth_header


# ---------------------------------------------------------- helpers
def _xlsx_bytes_from_rows(rows: list[dict]) -> bytes:
    """Build a fresh .xlsx with the canonical headers + given data rows.

    Each `rows` entry is a dict keyed by header name; missing keys
    become blank cells. This mirrors what an admin's filled-in
    template would look like over the wire.
    """
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=name)
    for ri, row in enumerate(rows, start=2):
        for col, name in enumerate(HEADERS, start=1):
            v = row.get(name)
            if v is not None:
                ws.cell(row=ri, column=col, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ok_single(stem="A real question?", topic="DU") -> dict:
    """A minimal valid single-choice row."""
    return {
        "stem": stem, "topic_code": topic,
        "difficulty": "easy", "question_type": "single_choice",
        "option_a_text": "yes",  "option_a_is_correct": "true",
        "option_b_text": "no",   "option_b_is_correct": "false",
    }


def _ok_multi(stem="Pick all", topic="DU") -> dict:
    return {
        "stem": stem, "topic_code": topic,
        "difficulty": "medium", "question_type": "multi_choice",
        "option_a_text": "x", "option_a_is_correct": "true",
        "option_b_text": "y", "option_b_is_correct": "true",
        "option_c_text": "z", "option_c_is_correct": "false",
    }


def _post_upload(client, headers, blob):
    return client.post(
        "/api/v1/admin/questions/bulk-upload",
        headers=headers,
        files={"file": ("upload.xlsx", blob,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ============================================================ template
def test_template_download_returns_valid_xlsx(client, admin):
    h = auth_header(client, admin.email)
    r = client.get("/api/v1/admin/questions/bulk-template", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "filename=" in r.headers.get("content-disposition", "")
    # Round-trip: open it back and check headers match.
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    headers_in_file = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
    for required in ("stem", "topic_code", "difficulty",
                      "option_a_text", "option_a_is_correct"):
        assert required in headers_in_file, f"template missing column {required!r}"


def test_template_includes_example_rows_for_self_documentation(client, admin):
    """Template ships with example rows so admins can see the expected
    shape without reading docs. Verify at least one example row is
    present and has the multi_choice question_type so admins discover
    that mode exists."""
    h = auth_header(client, admin.email)
    r = client.get("/api/v1/admin/questions/bulk-template", headers=h)
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    nonblank = [r for r in rows if r and r[0]]
    assert len(nonblank) >= 1, "template should have at least one example row"
    # qtype is column 4 (1-indexed)
    qtypes = [r[3] for r in nonblank if r[3]]
    assert "multi_choice" in qtypes, (
        "template should include a multi_choice example so admins "
        "discover the feature")


# ====================================================== happy path
def test_bulk_upload_creates_all_valid_rows(client, admin, db):
    h = auth_header(client, admin.email)
    blob = _xlsx_bytes_from_rows([
        _ok_single("Question one?", "DU"),
        _ok_single("Question two?", "BU"),
        _ok_multi("Pick all that apply",   "DP"),
    ])
    r = _post_upload(client, h, blob)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 3
    assert body["errors"] == []
    assert len(body["created_ids"]) == 3
    # Verify one row landed with all its fields populated.
    q = db.get(Question, body["created_ids"][0])
    assert q.stem == "Question one?"
    assert q.created_by == admin.id


# ============================================== per-row partial success
def test_bulk_upload_skips_bad_rows_keeps_good_ones(client, admin, db):
    """Three rows: good, bad (unknown topic), good. Should create 2
    and report 1 error referencing row 3."""
    h = auth_header(client, admin.email)
    blob = _xlsx_bytes_from_rows([
        _ok_single("good 1", "DU"),
        _ok_single("bad", "ZZ"),                  # unknown topic_code
        _ok_single("good 2", "BU"),
    ])
    r = _post_upload(client, h, blob)
    assert r.status_code == 200
    body = r.json()
    assert body["created"] == 2
    assert len(body["errors"]) == 1
    err = body["errors"][0]
    assert err["row"] == 3                        # 1=header, 2=row1, 3=row2
    assert "ZZ" in err["message"]


def test_bulk_upload_runs_same_validator_as_single_create(client, admin):
    """A multi-choice row with only 1 correct option must fail the
    same way `POST /admin/questions` would — proves the bulk path
    can't drift from the single path's rules."""
    h = auth_header(client, admin.email)
    bad_multi = {
        "stem": "Bad multi", "topic_code": "DU",
        "difficulty": "easy", "question_type": "multi_choice",
        "option_a_text": "x", "option_a_is_correct": "true",
        "option_b_text": "y", "option_b_is_correct": "false",
    }
    blob = _xlsx_bytes_from_rows([bad_multi])
    r = _post_upload(client, h, blob)
    body = r.json()
    assert body["created"] == 0
    assert len(body["errors"]) == 1
    assert "at least 2" in body["errors"][0]["message"].lower()


def test_bulk_upload_single_with_two_correct_fails_validation(client, admin):
    h = auth_header(client, admin.email)
    bad_single = {
        "stem": "Bad single", "topic_code": "DU",
        "difficulty": "easy", "question_type": "single_choice",
        "option_a_text": "x", "option_a_is_correct": "true",
        "option_b_text": "y", "option_b_is_correct": "true",
    }
    blob = _xlsx_bytes_from_rows([bad_single])
    body = _post_upload(client, h, blob).json()
    assert body["created"] == 0
    assert "exactly one" in body["errors"][0]["message"].lower()


def test_bulk_upload_missing_stem_caught_by_parser(client, admin):
    h = auth_header(client, admin.email)
    bad = _ok_single()
    bad.pop("stem")
    blob = _xlsx_bytes_from_rows([bad])
    body = _post_upload(client, h, blob).json()
    assert body["created"] == 0
    assert "stem" in body["errors"][0]["message"].lower()


def test_bulk_upload_only_one_option_fails(client, admin):
    h = auth_header(client, admin.email)
    bad = {
        "stem": "Lonely", "topic_code": "DU", "difficulty": "easy",
        "option_a_text": "only one", "option_a_is_correct": "true",
    }
    body = _post_upload(client, h, bad and _xlsx_bytes_from_rows([bad])).json()
    assert body["created"] == 0
    assert "2 options" in body["errors"][0]["message"]


# ============================================= file-shape guards
def test_bulk_upload_empty_file_rejected(client, admin):
    h = auth_header(client, admin.email)
    r = client.post("/api/v1/admin/questions/bulk-upload",
                     headers=h,
                     files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 422
    assert "empty" in r.json()["error"]["message"].lower()


def test_bulk_upload_oversized_file_rejected(client, admin, monkeypatch):
    """Patch the cap down to 100 bytes so we can trigger the guard
    without sending a real 5MB file in the test runner."""
    monkeypatch.setattr(questions_ep, "MAX_UPLOAD_BYTES", 100)
    h = auth_header(client, admin.email)
    blob = _xlsx_bytes_from_rows([_ok_single()])
    assert len(blob) > 100, "test fixture must exceed the patched cap"
    r = _post_upload(client, h, blob)
    assert r.status_code == 422
    assert "too large" in r.json()["error"]["message"].lower()


def test_bulk_upload_corrupt_xlsx_returns_clean_error(client, admin):
    h = auth_header(client, admin.email)
    r = client.post("/api/v1/admin/questions/bulk-upload",
                     headers=h,
                     files={"file": ("notxlsx.xlsx", b"this is not an xlsx",
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 0
    assert len(body["errors"]) == 1
    assert "could not open" in body["errors"][0]["message"].lower()


def test_bulk_upload_missing_headers_returns_clean_error(client, admin):
    """Old template or hand-rolled sheet missing required columns →
    clean error, not a 500."""
    h = auth_header(client, admin.email)
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="some_random_header")
    ws.cell(row=2, column=1, value="data")
    buf = io.BytesIO(); wb.save(buf)
    r = _post_upload(client, h, buf.getvalue())
    body = r.json()
    assert body["created"] == 0
    assert "missing required header" in body["errors"][0]["message"].lower()


# ================================================= blank rows ignored
def test_bulk_upload_skips_trailing_blank_rows(client, admin, db):
    """Admins often leave 10 trailing blank rows in their sheet — these
    must not be counted as errors."""
    h = auth_header(client, admin.email)
    blob = _xlsx_bytes_from_rows([
        _ok_single("only real question", "DU"),
        {},        # blank
        {},        # blank
        {},        # blank
    ])
    body = _post_upload(client, h, blob).json()
    assert body["created"] == 1
    assert body["errors"] == []


# =========================================================== audit log
def test_bulk_upload_writes_audit_entry(client, admin, db):
    h = auth_header(client, admin.email)
    blob = _xlsx_bytes_from_rows([_ok_single(), _ok_single("two", "BU")])
    _post_upload(client, h, blob)
    audit = (db.query(AuditLog)
             .filter_by(action="question.bulk_upload").first())
    assert audit is not None
    assert audit.metadata_json.get("created") == 2
    assert audit.metadata_json.get("filename") == "upload.xlsx"


# ========================================================== auth gate
def test_bulk_endpoints_require_admin(client, user):
    """Regular users get rejected at the /admin/* gate."""
    h = auth_header(client, user.email)
    r1 = client.get("/api/v1/admin/questions/bulk-template", headers=h)
    assert r1.status_code in (401, 403)
    r2 = _post_upload(client, h, _xlsx_bytes_from_rows([_ok_single()]))
    assert r2.status_code in (401, 403)
